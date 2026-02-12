from PySide6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                               QLabel, QPushButton, QTableWidget, QTableWidgetItem,
                               QComboBox, QSpinBox, QCheckBox, QHeaderView, QMessageBox,
                               QFileDialog)
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from frontend.styles import AppStyles
from backend.api.asistencias_api import AsistenciasAPI
from backend.api.cierres_api import CierresAPI
from datetime import datetime, date
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class AsistenciasWindow(QMainWindow):
    def __init__(self, user_email):
        super().__init__()
        self.user_email = user_email
        self.asistencias_api = AsistenciasAPI()
        self.cierres_api = CierresAPI()
        self.empleados_data = []
        self.asistencias_data = {}
        self.mes_actual = datetime.now().month
        self.anio_actual = datetime.now().year
        self.periodo_cerrado = False  # Bandera para controlar si está cerrado
        self.init_ui()
    
    def init_ui(self):
        """Inicializa la interfaz de usuario"""
        self.setWindowTitle("Registro de Asistencia")
        self.setMinimumSize(1400, 700)
        self.showMaximized()  # Abrir maximizada
        self.setStyleSheet(AppStyles.WINDOW_STYLE)
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout principal
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Título
        title = QLabel("REGISTRO DE ASISTENCIA")
        title.setStyleSheet("""
            QLabel {
                color: #FFFFFF;
                font-size: 24px;
                font-weight: bold;
                padding: 15px;
                background-color: rgba(255, 140, 0, 0.3);
                border-radius: 8px;
            }
        """)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)
        
        # Panel de controles superiores
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(15)
        
        # Label Mes
        mes_label = QLabel("Mes:")
        mes_label.setStyleSheet(AppStyles.LABEL_STYLE)
        controls_layout.addWidget(mes_label)
        
        # ComboBox Mes
        self.mes_combo = QComboBox()
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        self.mes_combo.addItems(meses)
        self.mes_combo.setCurrentIndex(self.mes_actual - 1)
        self.mes_combo.setStyleSheet("""
            QComboBox {
                background-color: #F0F8FF;
                color: #000000;
                border: 2px solid #B0C4DE;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 14px;
                min-width: 120px;
            }
        """)
        controls_layout.addWidget(self.mes_combo)
        
        # Label Año
        anio_label = QLabel("Año:")
        anio_label.setStyleSheet(AppStyles.LABEL_STYLE)
        controls_layout.addWidget(anio_label)
        
        # SpinBox Año
        self.anio_spin = QSpinBox()
        self.anio_spin.setRange(2020, 2050)
        self.anio_spin.setValue(self.anio_actual)
        self.anio_spin.setStyleSheet("""
            QSpinBox {
                background-color: #F0F8FF;
                color: #000000;
                border: 2px solid #B0C4DE;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 14px;
                min-width: 80px;
            }
        """)
        controls_layout.addWidget(self.anio_spin)
        
        # Botón Cargar
        self.btn_cargar = QPushButton("CARGAR")
        self.btn_cargar.setStyleSheet(AppStyles.BUTTON_STYLE)
        self.btn_cargar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cargar.clicked.connect(self.cargar_asistencias)
        controls_layout.addWidget(self.btn_cargar)
        
        # Botón Exportar a Excel
        self.btn_excel = QPushButton("EXPORTAR A EXCEL")
        self.btn_excel.setStyleSheet(AppStyles.BUTTON_STYLE)
        self.btn_excel.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_excel.clicked.connect(self.exportar_excel)
        self.btn_excel.setEnabled(False)  # Deshabilitado hasta cargar datos
        controls_layout.addWidget(self.btn_excel)
        
        controls_layout.addStretch()
        
        # Botón Cerrar
        self.btn_cerrar = QPushButton("CERRAR")
        self.btn_cerrar.setStyleSheet("""
            QPushButton {
                background-color: #DC3545;
                color: #FFFFFF;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #C82333;
            }
        """)
        self.btn_cerrar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cerrar.clicked.connect(self.close)
        controls_layout.addWidget(self.btn_cerrar)
        
        main_layout.addLayout(controls_layout)
        
        # Tabla de asistencias
        self.table = QTableWidget()
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: #F0F8FF;
                border: 2px solid #B0C4DE;
                border-radius: 8px;
                gridline-color: #B0C4DE;
            }
            QTableWidget::item {
                padding: 5px;
                color: #000000;
            }
            QHeaderView::section {
                background-color: #4682B4;
                color: #FFFFFF;
                padding: 8px;
                border: 1px solid #B0C4DE;
                font-weight: bold;
                font-size: 12px;
            }
        """)
        
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        main_layout.addWidget(self.table)
        
        # Cargar datos iniciales
        self.cargar_asistencias()
    
    def cargar_asistencias(self):
        """Carga los empleados vigentes y sus asistencias del mes seleccionado"""
        mes = self.mes_combo.currentIndex() + 1
        anio = self.anio_spin.value()
        
        # Verificar cierre del período
        periodo = f"{anio}{mes:02d}"  # Formato: YYYYMM (ej: 202502)
        
        try:
            # Verificar si el período está cerrado
            resultado_cierre = self.cierres_api.verificar_cierre(periodo)
            
            if not resultado_cierre["success"]:
                QMessageBox.warning(self, "Error", resultado_cierre["message"])
                return
            
            self.periodo_cerrado = resultado_cierre.get("cerrado", False)
            
            # Si está cerrado, mostrar advertencia
            if self.periodo_cerrado:
                QMessageBox.warning(
                    self,
                    "Período Cerrado",
                    f"El período {mes:02d}/{anio} está CERRADO.\n\n"
                    f"{resultado_cierre.get('mensaje', '')}\n\n"
                    "Solo podrá visualizar los datos, no podrá realizar cambios."
                )
            
            # Continuar cargando datos
            result = self.asistencias_api.get_asistencias_mes(mes, anio)
            
            if not result["success"]:
                QMessageBox.warning(self, "Error", result["message"])
                return
            
            self.empleados_data = result["data"]["empleados"]
            asistencias_list = result["data"]["asistencias"]
            
            # Debug: Mostrar cuántas asistencias se cargaron
            print(f"Empleados cargados: {len(self.empleados_data)}")
            print(f"Asistencias cargadas: {len(asistencias_list)}")
            print(f"Período cerrado: {self.periodo_cerrado}")
            
            # Convertir lista de asistencias a diccionario
            self.asistencias_data = {}
            for asist in asistencias_list:
                # Convertir fecha a string si viene como objeto date
                if isinstance(asist['fecha'], date):
                    fecha_str = asist['fecha'].strftime("%Y-%m-%d")
                else:
                    fecha_str = str(asist['fecha'])
                
                key = f"{asist['asistencia_empleado_codigo']}_{fecha_str}"
                self.asistencias_data[key] = asist['estado']
                print(f"Cargada asistencia: {key} = {asist['estado']}")  # Debug
            
            self.crear_tabla(mes, anio)
            
            # Habilitar botón de Excel si hay datos
            self.btn_excel.setEnabled(len(self.empleados_data) > 0)
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar datos: {str(e)}")
            import traceback
            print(traceback.format_exc())  # Debug completo
    
    def crear_tabla(self, mes, anio):
        """Crea la tabla con los días del mes y empleados"""
        # Obtener días del mes
        dias_mes = calendar.monthrange(anio, mes)[1]
        
        # Configurar columnas: Código + Nombre + días del mes + Total (SIN ID)
        num_columnas = 2 + dias_mes + 1
        self.table.setColumnCount(num_columnas)
        self.table.setRowCount(len(self.empleados_data))
        
        # Headers (SIN ID)
        headers = ["Código", "Empleado"]
        
        # Agregar días del mes con día de la semana (abreviado a 1 letra)
        for dia in range(1, dias_mes + 1):
            fecha = date(anio, mes, dia)
            dia_semana = fecha.strftime("%a")[0]  # Solo primera letra: L, M, M, J, V, S, D
            headers.append(f"{dia}")  # Solo el número, sin día de semana en header
        
        headers.append("T")  # Total
        self.table.setHorizontalHeaderLabels(headers)
        
        # Ajustar ancho de columnas (mínimo absoluto final)
        self.table.setColumnWidth(0, 30)   # Código
        self.table.setColumnWidth(1, 70)   # Nombre
        
        for col in range(2, num_columnas - 1):  # Días
            self.table.setColumnWidth(col, 14)  # Mínimo absoluto final
        
        self.table.setColumnWidth(num_columnas - 1, 20)  # Total
        
        # Ajustar altura de las filas
        for row in range(len(self.empleados_data)):
            self.table.setRowHeight(row, 18)
        
        # Llenar filas con empleados
        for row, empleado in enumerate(self.empleados_data):
            # Código
            codigo_item = QTableWidgetItem(str(empleado['empleado_codigo']))
            codigo_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 0, codigo_item)
            
            # Nombre
            nombre_item = QTableWidgetItem(empleado['empleado_nombres'])
            self.table.setItem(row, 1, nombre_item)
            
            # Checkboxes para cada día
            total = 0
            for dia in range(1, dias_mes + 1):
                col = dia + 1  # +1 porque las primeras 2 columnas son código y nombre
                fecha = date(anio, mes, dia)
                fecha_str = fecha.strftime("%Y-%m-%d")
                
                # Crear item de celda primero (sin color de fondo por ahora)
                cell_item = QTableWidgetItem("")
                cell_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, col, cell_item)
                
                # Crear checkbox sin estilos complicados
                checkbox = QCheckBox()
                
                # Si el período está cerrado, deshabilitar el checkbox
                if self.periodo_cerrado:
                    checkbox.setEnabled(False)
                
                # Verificar si existe asistencia
                key = f"{empleado['empleado_codigo']}_{fecha_str}"
                if key in self.asistencias_data and self.asistencias_data[key] == 'P':
                    checkbox.setChecked(True)
                    total += 1
                
                # Conectar señal solo si no está cerrado
                if not self.periodo_cerrado:
                    checkbox.stateChanged.connect(
                        lambda state, r=row, d=dia, emp=empleado['empleado_codigo'], f=fecha_str: 
                        self.on_checkbox_changed(r, d, emp, f, state)
                    )
                
                # Agregar checkbox a la celda centrado (sin margin/padding)
                cell_widget = QWidget()
                cell_layout = QHBoxLayout(cell_widget)
                cell_layout.addWidget(checkbox)
                cell_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                cell_layout.setContentsMargins(0, 0, 0, 0)
                cell_layout.setSpacing(0)
                cell_widget.setMaximumWidth(20)  # Forzar ancho mínimo
                
                # Marcar visualmente los fines de semana en el widget contenedor
                dia_semana = fecha.weekday()
                if dia_semana == 5:  # Sábado
                    cell_widget.setStyleSheet("background-color: #ADD8E6;")  # Azul claro
                elif dia_semana == 6:  # Domingo
                    cell_widget.setStyleSheet("background-color: #FFB6C1;")  # Rosa claro
                
                self.table.setCellWidget(row, col, cell_widget)
            
            # Total
            total_item = QTableWidgetItem(str(total))
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            total_item.setBackground(QColor(144, 238, 144))  # Verde claro
            font = total_item.font()
            font.setBold(True)
            total_item.setFont(font)
            self.table.setItem(row, num_columnas - 1, total_item)
    
    def on_checkbox_changed(self, row, dia, empleado_codigo, fecha_str, state):
        """Maneja el cambio de estado de un checkbox"""
        try:
            if state == Qt.CheckState.Checked.value:
                # Marcar como presente
                result = self.asistencias_api.registrar_asistencia(
                    empleado_codigo,
                    fecha_str,
                    'P',
                    self.user_email
                )
                
                if result["success"]:
                    # Actualizar total
                    self.actualizar_total(row)
                else:
                    QMessageBox.warning(self, "Error", result["message"])
            else:
                # Eliminar asistencia (marcar como ausente)
                result = self.asistencias_api.eliminar_asistencia(
                    empleado_codigo,
                    fecha_str
                )
                
                if result["success"]:
                    # Actualizar total
                    self.actualizar_total(row)
                else:
                    QMessageBox.warning(self, "Error", result["message"])
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al registrar: {str(e)}")
    
    def actualizar_total(self, row):
        """Recalcula y actualiza el total de una fila"""
        total = 0
        num_columnas = self.table.columnCount()
        
        # Contar checkboxes marcados (desde columna 2 hasta la penúltima)
        for col in range(2, num_columnas - 1):
            widget = self.table.cellWidget(row, col)
            if widget:
                checkbox = widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    total += 1
        
        # Actualizar celda de total
        total_item = self.table.item(row, num_columnas - 1)
        if total_item:
            total_item.setText(str(total))
    
    def exportar_excel(self):
        """Exporta la tabla de asistencias a Excel"""
        if not self.empleados_data:
            QMessageBox.warning(self, "Advertencia", "No hay datos para exportar")
            return
        
        mes = self.mes_combo.currentIndex() + 1
        anio = self.anio_spin.value()
        nombre_mes = self.mes_combo.currentText()
        
        # Diálogo para guardar archivo
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar archivo Excel",
            f"asistencias_{nombre_mes}_{anio}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Asistencias {nombre_mes}"
            
            # Obtener días del mes
            dias_mes = calendar.monthrange(anio, mes)[1]
            
            # Estilos
            header_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            presente_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            sabado_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            domingo_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título principal
            ws.merge_cells('A1:C1')
            titulo = ws['A1']
            titulo.value = f"REGISTRO DE ASISTENCIA - {nombre_mes.upper()} {anio}"
            titulo.font = Font(bold=True, size=14, color="4682B4")
            titulo.alignment = Alignment(horizontal="center", vertical="center")
            
            # Encabezados (fila 3)
            row_idx = 3
            headers = ["Código", "Empleado"]
            
            # Agregar días
            for dia in range(1, dias_mes + 1):
                headers.append(str(dia))
            
            headers.append("Total")
            
            # Escribir encabezados
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 10  # Código
            ws.column_dimensions['B'].width = 30  # Empleado
            
            # Columnas de días
            for col_idx in range(3, 3 + dias_mes):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 4
            
            # Columna Total
            total_col_letter = openpyxl.utils.get_column_letter(3 + dias_mes)
            ws.column_dimensions[total_col_letter].width = 6
            
            # Datos de empleados
            row_idx = 4
            for empleado in self.empleados_data:
                # Código
                cell = ws.cell(row=row_idx, column=1, value=empleado['empleado_codigo'])
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                
                # Nombre
                cell = ws.cell(row=row_idx, column=2, value=empleado['empleado_nombres'])
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                
                # Asistencias por día
                total_presente = 0
                for dia in range(1, dias_mes + 1):
                    col_idx = dia + 2  # +2 porque las primeras 2 son código y nombre
                    fecha = date(anio, mes, dia)
                    fecha_str = fecha.strftime("%Y-%m-%d")
                    
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Verificar asistencia
                    key = f"{empleado['empleado_codigo']}_{fecha_str}"
                    if key in self.asistencias_data and self.asistencias_data[key] == 'P':
                        cell.value = "✓"
                        cell.fill = presente_fill
                        cell.font = Font(bold=True, size=12, color="006400")
                        total_presente += 1
                    else:
                        cell.value = ""
                    
                    # Colorear fines de semana
                    dia_semana = fecha.weekday()
                    if dia_semana == 5:  # Sábado
                        if not cell.fill or cell.fill.start_color.rgb != "0090EE90":
                            cell.fill = sabado_fill
                    elif dia_semana == 6:  # Domingo
                        if not cell.fill or cell.fill.start_color.rgb != "0090EE90":
                            cell.fill = domingo_fill
                
                # Total
                cell = ws.cell(row=row_idx, column=3 + dias_mes, value=total_presente)
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                
                row_idx += 1
            
            # Leyenda
            row_idx += 2
            ws.cell(row=row_idx, column=1, value="Leyenda:")
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            
            row_idx += 1
            cell = ws.cell(row=row_idx, column=1, value="✓")
            cell.fill = presente_fill
            cell.font = Font(bold=True, size=12, color="006400")
            cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value="= Presente")
            
            row_idx += 1
            cell = ws.cell(row=row_idx, column=1)
            cell.fill = sabado_fill
            cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value="= Sábado")
            
            row_idx += 1
            cell = ws.cell(row=row_idx, column=1)
            cell.fill = domingo_fill
            cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value="= Domingo")
            
            # Pie de página
            row_idx += 2
            ws.cell(row=row_idx, column=1, 
                   value=f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            ws.cell(row=row_idx, column=1).font = Font(italic=True, size=9)
            
            # Guardar archivo
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "Éxito",
                f"Asistencias exportadas exitosamente a:\n{file_path}"
            )
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")
            import traceback
            print(traceback.format_exc())