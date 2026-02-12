from PySide6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                               QLabel, QPushButton, QTableWidget, QTableWidgetItem,
                               QDialog, QLineEdit, QMessageBox, QHeaderView, QComboBox,
                               QFileDialog)
from PySide6.QtCore import Qt
from PySide6.QtGui import QIcon
from frontend.styles import AppStyles
from backend.api.empleados_api import EmpleadosAPI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class EmpleadosWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.empleados_api = EmpleadosAPI()
        self.empleados_data = []
        self.init_ui()
        self.load_empleados()
    
    def init_ui(self):
        """Inicializa la interfaz de usuario"""
        self.setWindowTitle("Mantenimiento de Empleados")
        self.setMinimumSize(1000, 600)
        self.setStyleSheet(AppStyles.WINDOW_STYLE)
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout principal
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 20, 30, 20)
        
        # Título
        title = QLabel("MANTENIMIENTO DE EMPLEADOS")
        title.setStyleSheet("""
            QLabel {
                color: #FFFFFF;
                font-size: 24px;
                font-weight: bold;
                padding: 15px;
                background-color: rgba(255, 140, 0, 0.3);
                border-radius: 8px;
            }
        """)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)
        
        # Botones superiores
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)
        
        # Botón Agregar
        self.btn_agregar = QPushButton("AGREGAR")
        self.btn_agregar.setStyleSheet(AppStyles.BUTTON_STYLE)
        self.btn_agregar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_agregar.clicked.connect(self.agregar_empleado)
        buttons_layout.addWidget(self.btn_agregar)
        
        # Botón Modificar
        self.btn_modificar = QPushButton("MODIFICAR")
        self.btn_modificar.setStyleSheet(AppStyles.BUTTON_STYLE)
        self.btn_modificar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_modificar.clicked.connect(self.modificar_empleado)
        self.btn_modificar.setEnabled(False)
        buttons_layout.addWidget(self.btn_modificar)
        
        # Botón Excel
        self.btn_excel = QPushButton("EXPORTAR A EXCEL")
        self.btn_excel.setStyleSheet(AppStyles.BUTTON_STYLE)
        self.btn_excel.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_excel.clicked.connect(self.exportar_excel)
        buttons_layout.addWidget(self.btn_excel)
        
        # Botón Cancelar
        self.btn_cancelar = QPushButton("CANCELAR")
        self.btn_cancelar.setStyleSheet("""
            QPushButton {
                background-color: #DC3545;
                color: #FFFFFF;
                border: none;
                border-radius: 8px;
                padding: 12px 24px;
                font-size: 14px;
                font-weight: bold;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #C82333;
            }
        """)
        self.btn_cancelar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cancelar.clicked.connect(self.close)
        buttons_layout.addWidget(self.btn_cancelar)
        
        buttons_layout.addStretch()
        main_layout.addLayout(buttons_layout)
        
        # Tabla de empleados
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels([
            "ID", "Código", "Nombres", "Email", "Estado"
        ])
        
        # Estilos de la tabla
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: #F0F8FF;
                border: 2px solid #B0C4DE;
                border-radius: 8px;
                gridline-color: #B0C4DE;
            }
            QTableWidget::item {
                padding: 8px;
                color: #000000;
            }
            QTableWidget::item:selected {
                background-color: #FF8C00;
                color: #FFFFFF;
            }
            QHeaderView::section {
                background-color: #4682B4;
                color: #FFFFFF;
                padding: 10px;
                border: 1px solid #B0C4DE;
                font-weight: bold;
            }
        """)
        
        # Configurar tabla
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        
        main_layout.addWidget(self.table)
    
    def load_empleados(self):
        """Carga los empleados desde la API"""
        try:
            result = self.empleados_api.get_all_empleados()
            
            if result["success"]:
                self.empleados_data = result["data"]
                self.populate_table()
            else:
                QMessageBox.warning(self, "Error", result["message"])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar empleados: {str(e)}")
    
    def populate_table(self):
        """Llena la tabla con los datos de empleados"""
        self.table.setRowCount(len(self.empleados_data))
        
        for row, empleado in enumerate(self.empleados_data):
            self.table.setItem(row, 0, QTableWidgetItem(str(empleado['empleado_id'])))
            self.table.setItem(row, 1, QTableWidgetItem(str(empleado['empleado_codigo'])))
            self.table.setItem(row, 2, QTableWidgetItem(empleado['empleado_nombres']))
            self.table.setItem(row, 3, QTableWidgetItem(empleado['empleado_email'] or ""))
            
            # Estado con formato
            estado_text = "Vigente" if empleado['empleado_estado'] == 'V' else "Inactivo"
            self.table.setItem(row, 4, QTableWidgetItem(estado_text))
    
    def on_selection_changed(self):
        """Habilita/deshabilita el botón modificar según la selección"""
        self.btn_modificar.setEnabled(len(self.table.selectedItems()) > 0)
    
    def agregar_empleado(self):
        """Abre el diálogo para agregar un nuevo empleado"""
        dialog = AgregarEmpleadoDialog(self, self.empleados_api)
        if dialog.exec():
            self.load_empleados()
    
    def modificar_empleado(self):
        """Abre el diálogo para modificar el empleado seleccionado"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            return
        
        empleado = self.empleados_data[selected_row]
        dialog = ModificarEmpleadoDialog(self, self.empleados_api, empleado)
        if dialog.exec():
            self.load_empleados()
    
    def exportar_excel(self):
        """Exporta la tabla de empleados a Excel"""
        if not self.empleados_data:
            QMessageBox.warning(self, "Advertencia", "No hay datos para exportar")
            return
        
        # Diálogo para guardar archivo
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar archivo Excel",
            "empleados.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Empleados"
            
            # Estilos
            header_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            # Encabezados
            headers = ["ID", "Código", "Nombres", "Email", "Estado"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos
            for row_idx, empleado in enumerate(self.empleados_data, 2):
                ws.cell(row=row_idx, column=1, value=empleado['empleado_id'])
                ws.cell(row=row_idx, column=2, value=empleado['empleado_codigo'])
                ws.cell(row=row_idx, column=3, value=empleado['empleado_nombres'])
                ws.cell(row=row_idx, column=4, value=empleado['empleado_email'] or "")
                estado = "Vigente" if empleado['empleado_estado'] == 'V' else "Inactivo"
                ws.cell(row=row_idx, column=5, value=estado)
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 12
            
            # Guardar archivo
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "Éxito",
                f"Datos exportados exitosamente a:\n{file_path}"
            )
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")


class AgregarEmpleadoDialog(QDialog):
    def __init__(self, parent, api):
        super().__init__(parent)
        self.api = api
        self.init_ui()
    
    def init_ui(self):
        """Inicializa el diálogo de agregar empleado"""
        self.setWindowTitle("Agregar Empleado")
        self.setFixedSize(450, 300)
        self.setStyleSheet(AppStyles.WINDOW_STYLE)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(30, 20, 30, 20)
        
        # Título
        title = QLabel("Nuevo Empleado")
        title.setStyleSheet("""
            QLabel {
                color: #FFFFFF;
                font-size: 18px;
                font-weight: bold;
                padding: 10px;
            }
        """)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # Campo Código
        codigo_label = QLabel("Código:")
        codigo_label.setStyleSheet(AppStyles.LABEL_STYLE)
        layout.addWidget(codigo_label)
        
        self.codigo_input = QLineEdit()
        self.codigo_input.setPlaceholderText("Ej: 1001")
        self.codigo_input.setStyleSheet(AppStyles.INPUT_STYLE)
        layout.addWidget(self.codigo_input)
        
        # Campo Nombres
        nombres_label = QLabel("Nombres:")
        nombres_label.setStyleSheet(AppStyles.LABEL_STYLE)
        layout.addWidget(nombres_label)
        
        self.nombres_input = QLineEdit()
        self.nombres_input.setPlaceholderText("Nombre completo del empleado")
        self.nombres_input.setStyleSheet(AppStyles.INPUT_STYLE)
        layout.addWidget(self.nombres_input)
        
        # Mensaje de error
        self.error_label = QLabel("")
        self.error_label.setStyleSheet(AppStyles.ERROR_LABEL_STYLE)
        self.error_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.error_label.setWordWrap(True)
        layout.addWidget(self.error_label)
        
        # Botones
        buttons_layout = QHBoxLayout()
        
        btn_guardar = QPushButton("GUARDAR")
        btn_guardar.setStyleSheet(AppStyles.BUTTON_STYLE)
        btn_guardar.clicked.connect(self.guardar)
        buttons_layout.addWidget(btn_guardar)
        
        btn_cancelar = QPushButton("CANCELAR")
        btn_cancelar.setStyleSheet(AppStyles.BUTTON_STYLE)
        btn_cancelar.clicked.connect(self.reject)
        buttons_layout.addWidget(btn_cancelar)
        
        layout.addLayout(buttons_layout)
    
    def guardar(self):
        """Guarda el nuevo empleado"""
        self.error_label.setText("")
        
        codigo_text = self.codigo_input.text().strip()
        nombres = self.nombres_input.text().strip()
        
        # Validaciones
        if not codigo_text:
            self.error_label.setText("El código es requerido")
            return
        
        try:
            codigo = int(codigo_text)
        except ValueError:
            self.error_label.setText("El código debe ser un número")
            return
        
        if not nombres:
            self.error_label.setText("El nombre es requerido")
            return
        
        # Verificar si el código ya existe
        check_result = self.api.check_codigo_exists(codigo)
        if check_result.get("exists"):
            self.error_label.setText(f"El código {codigo} ya existe")
            return
        
        # Crear empleado
        result = self.api.create_empleado({
            'empleado_codigo': codigo,
            'empleado_nombres': nombres
        })
        
        if result["success"]:
            QMessageBox.information(self, "Éxito", "Empleado agregado exitosamente")
            self.accept()
        else:
            self.error_label.setText(result["message"])


class ModificarEmpleadoDialog(QDialog):
    def __init__(self, parent, api, empleado):
        super().__init__(parent)
        self.api = api
        self.empleado = empleado
        self.init_ui()
    
    def init_ui(self):
        """Inicializa el diálogo de modificar empleado"""
        self.setWindowTitle("Modificar Empleado")
        self.setFixedSize(450, 350)
        self.setStyleSheet(AppStyles.WINDOW_STYLE)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(30, 20, 30, 20)
        
        # Título
        title = QLabel(f"Modificar Empleado - Código: {self.empleado['empleado_codigo']}")
        title.setStyleSheet("""
            QLabel {
                color: #FFFFFF;
                font-size: 16px;
                font-weight: bold;
                padding: 10px;
            }
        """)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # Campo Nombres
        nombres_label = QLabel("Nombres:")
        nombres_label.setStyleSheet(AppStyles.LABEL_STYLE)
        layout.addWidget(nombres_label)
        
        self.nombres_input = QLineEdit()
        self.nombres_input.setText(self.empleado['empleado_nombres'])
        self.nombres_input.setStyleSheet(AppStyles.INPUT_STYLE)
        layout.addWidget(self.nombres_input)
        
        # Campo Estado
        estado_label = QLabel("Estado:")
        estado_label.setStyleSheet(AppStyles.LABEL_STYLE)
        layout.addWidget(estado_label)
        
        self.estado_combo = QComboBox()
        self.estado_combo.addItem("Vigente", "V")
        self.estado_combo.addItem("Inactivo", "I")
        self.estado_combo.setStyleSheet("""
            QComboBox {
                background-color: #F0F8FF;
                color: #000000;
                border: 2px solid #B0C4DE;
                border-radius: 6px;
                padding: 10px 12px;
                font-size: 14px;
                min-height: 35px;
            }
            QComboBox:focus {
                border: 2px solid #FF8C00;
            }
            QComboBox::drop-down {
                border: none;
            }
        """)
        
        # Establecer estado actual
        if self.empleado['empleado_estado'] == 'V':
            self.estado_combo.setCurrentIndex(0)
        else:
            self.estado_combo.setCurrentIndex(1)
        
        layout.addWidget(self.estado_combo)
        
        # Mensaje de error
        self.error_label = QLabel("")
        self.error_label.setStyleSheet(AppStyles.ERROR_LABEL_STYLE)
        self.error_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.error_label.setWordWrap(True)
        layout.addWidget(self.error_label)
        
        # Botones
        buttons_layout = QHBoxLayout()
        
        btn_guardar = QPushButton("GUARDAR")
        btn_guardar.setStyleSheet(AppStyles.BUTTON_STYLE)
        btn_guardar.clicked.connect(self.guardar)
        buttons_layout.addWidget(btn_guardar)
        
        btn_cancelar = QPushButton("CANCELAR")
        btn_cancelar.setStyleSheet(AppStyles.BUTTON_STYLE)
        btn_cancelar.clicked.connect(self.reject)
        buttons_layout.addWidget(btn_cancelar)
        
        layout.addLayout(buttons_layout)
    
    def guardar(self):
        """Guarda los cambios del empleado"""
        self.error_label.setText("")
        
        nombres = self.nombres_input.text().strip()
        estado = self.estado_combo.currentData()
        
        # Validaciones
        if not nombres:
            self.error_label.setText("El nombre es requerido")
            return
        
        # Actualizar empleado
        result = self.api.update_empleado(
            self.empleado['empleado_id'],
            {
                'empleado_nombres': nombres,
                'empleado_estado': estado
            }
        )
        
        if result["success"]:
            QMessageBox.information(self, "Éxito", "Empleado actualizado exitosamente")
            self.accept()
        else:
            self.error_label.setText(result["message"])